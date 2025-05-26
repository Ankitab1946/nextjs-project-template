"""Module for generating various comparison reports"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
from datetime import datetime
from ydata_profiling import ProfileReport
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from config import EXCEL_STYLES, REPORTS_DIR

class ReportGenerator:
    @staticmethod
    def generate_regression_report(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                                 column_mapping: Dict[str, str], output_path: str) -> None:
        """Generate regression report with multiple checks"""
        
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # Generate Aggregation Check
        ReportGenerator._generate_aggregation_check(source_df, target_df, column_mapping, writer)
        
        # Generate Count Check
        ReportGenerator._generate_count_check(source_df, target_df, writer)
        
        # Generate Distinct Check
        ReportGenerator._generate_distinct_check(source_df, target_df, column_mapping, writer)
        
        writer.save()

    @staticmethod
    def _generate_aggregation_check(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                                  column_mapping: Dict[str, str], writer: pd.ExcelWriter) -> None:
        """Generate aggregation check for numeric columns"""
        
        agg_data = []
        for source_col, target_col in column_mapping.items():
            # Check if columns are numeric
            if pd.api.types.is_numeric_dtype(source_df[source_col]) and pd.api.types.is_numeric_dtype(target_df[target_col]):
                source_sum = source_df[source_col].sum()
                target_sum = target_df[target_col].sum()
                result = 'PASS' if np.isclose(source_sum, target_sum, rtol=1e-05) else 'FAIL'
                
                agg_data.append({
                    'Source Column': source_col,
                    'Target Column': target_col,
                    'Source Sum': source_sum,
                    'Target Sum': target_sum,
                    'Result': result
                })
        
        if agg_data:
            df_agg = pd.DataFrame(agg_data)
            df_agg.to_excel(writer, sheet_name='AggregationCheck', index=False)
            
            # Apply conditional formatting
            worksheet = writer.sheets['AggregationCheck']
            for idx, row in enumerate(df_agg['Result'], start=2):  # start=2 to skip header
                cell = worksheet[f'E{idx}']
                style = EXCEL_STYLES['PASS'] if row == 'PASS' else EXCEL_STYLES['FAIL']
                cell.fill = PatternFill(start_color=style['fill']['fgColor'], end_color=style['fill']['fgColor'], fill_type='solid')
                cell.font = Font(color=style['font']['color'])

    @staticmethod
    def _generate_count_check(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                            writer: pd.ExcelWriter) -> None:
        """Generate count check report"""
        
        count_data = {
            'Source Count': len(source_df),
            'Target Count': len(target_df),
            'Result': 'PASS' if len(source_df) == len(target_df) else 'FAIL'
        }
        
        df_count = pd.DataFrame([count_data])
        df_count.to_excel(writer, sheet_name='CountCheck', index=False)
        
        # Apply conditional formatting
        worksheet = writer.sheets['CountCheck']
        cell = worksheet['C2']  # Result column, first data row
        style = EXCEL_STYLES['PASS'] if count_data['Result'] == 'PASS' else EXCEL_STYLES['FAIL']
        cell.fill = PatternFill(start_color=style['fill']['fgColor'], end_color=style['fill']['fgColor'], fill_type='solid')
        cell.font = Font(color=style['font']['color'])

    @staticmethod
    def _generate_distinct_check(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                               column_mapping: Dict[str, str], writer: pd.ExcelWriter) -> None:
        """Generate distinct value check for non-numeric columns"""
        
        distinct_data = []
        for source_col, target_col in column_mapping.items():
            # Check if columns are non-numeric
            if not pd.api.types.is_numeric_dtype(source_df[source_col]):
                source_distinct = source_df[source_col].nunique()
                target_distinct = target_df[target_col].nunique()
                
                source_values = set(source_df[source_col].dropna().unique())
                target_values = set(target_df[target_col].dropna().unique())
                
                values_match = source_values == target_values
                count_match = source_distinct == target_distinct
                result = 'PASS' if values_match and count_match else 'FAIL'
                
                distinct_data.append({
                    'Source Column': source_col,
                    'Target Column': target_col,
                    'Source Distinct Count': source_distinct,
                    'Target Distinct Count': target_distinct,
                    'Source Values': ', '.join(map(str, sorted(source_values))),
                    'Target Values': ', '.join(map(str, sorted(target_values))),
                    'Result': result
                })
        
        if distinct_data:
            df_distinct = pd.DataFrame(distinct_data)
            df_distinct.to_excel(writer, sheet_name='DistinctCheck', index=False)
            
            # Apply conditional formatting
            worksheet = writer.sheets['DistinctCheck']
            for idx, row in enumerate(df_distinct['Result'], start=2):  # start=2 to skip header
                cell = worksheet[f'G{idx}']
                style = EXCEL_STYLES['PASS'] if row == 'PASS' else EXCEL_STYLES['FAIL']
                cell.fill = PatternFill(start_color=style['fill']['fgColor'], end_color=style['fill']['fgColor'], fill_type='solid')
                cell.font = Font(color=style['font']['color'])

    @staticmethod
    def generate_side_by_side_report(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                                   column_mapping: Dict[str, str], join_keys: List[Tuple[str, str]], 
                                   output_path: str) -> None:
        """Generate side by side comparison report"""
        
        # Prepare DataFrames for comparison
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        # Merge DataFrames based on join keys
        if join_keys:
            source_join_cols = [key[0] for key in join_keys]
            target_join_cols = [key[1] for key in join_keys]
            
            merged_df = pd.merge(
                source_df[source_cols], 
                target_df[target_cols],
                left_on=source_join_cols,
                right_on=target_join_cols,
                how='outer',
                indicator=True
            )
            
            # Find differences
            diff_mask = merged_df['_merge'] != 'both'
            diff_df = merged_df[diff_mask].copy()
            
            if not diff_df.empty:
                # Rename columns to show source/target
                diff_df.columns = [f'Source_{col}' if col in source_cols else f'Target_{col}' 
                                 for col in diff_df.columns]
                
                # Save to Excel
                diff_df.to_excel(output_path, index=False)
                return True
            else:
                # Create empty DataFrame with message
                pd.DataFrame({'Message': ['No differences found']}).to_excel(output_path, index=False)
                return False
        else:
            # If no join keys, compare row by row
            comparison_df = pd.DataFrame()
            for s_col, t_col in column_mapping.items():
                comparison_df[f'Source_{s_col}'] = source_df[s_col]
                comparison_df[f'Target_{t_col}'] = target_df[t_col]
            
            comparison_df.to_excel(output_path, index=False)
            return True

    @staticmethod
    def generate_profile_reports(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                               timestamp: str) -> Dict[str, str]:
        """Generate Y-Data profiling reports"""
        
        reports = {}
        
        # Create reports directory
        os.makedirs(REPORTS_DIR, exist_ok=True)
        
        # Generate source profile
        source_profile = ProfileReport(
            source_df,
            title="Source Data Profile Report",
            minimal=False,
            correlations={"cramers": True},
            vars={"num": {"low_categorical_threshold": 0}}
        )
        source_path = os.path.join(REPORTS_DIR, f"SourceProfile_{timestamp}.html")
        source_profile.to_file(source_path)
        reports['source'] = source_path
        
        # Generate target profile
        target_profile = ProfileReport(
            target_df,
            title="Target Data Profile Report",
            minimal=False,
            correlations={"cramers": True},
            vars={"num": {"low_categorical_threshold": 0}}
        )
        target_path = os.path.join(REPORTS_DIR, f"TargetProfile_{timestamp}.html")
        target_profile.to_file(target_path)
        reports['target'] = target_path
        
        # Generate comparison profile
        comparison_df = pd.DataFrame()
        for col in source_df.columns:
            comparison_df[f'Source_{col}'] = source_df[col]
        for col in target_df.columns:
            comparison_df[f'Target_{col}'] = target_df[col]
            
        comparison_profile = ProfileReport(
            comparison_df,
            title="Source vs Target Comparison Profile",
            minimal=False,
            correlations={"cramers": True},
            vars={"num": {"low_categorical_threshold": 0}}
        )
        comparison_path = os.path.join(REPORTS_DIR, f"ComparisonProfile_{timestamp}.html")
        comparison_profile.to_file(comparison_path)
        reports['comparison'] = comparison_path
        
        return reports

    @staticmethod
    def generate_comparison_report(source_df: pd.DataFrame, target_df: pd.DataFrame,
                               column_mapping: Dict[str, str], join_keys: List[str],
                               output_path: str) -> None:
        """Generate pandas-based comparison report"""
        
        # Prepare DataFrames with mapped columns
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        if join_keys:
            source_join_cols = [key[0] for key in join_keys]
            target_join_cols = [key[1] for key in join_keys]
        else:
            # If no join keys, use index
            source_df = source_df.copy()
            target_df = target_df.copy()
            source_df['_index'] = range(len(source_df))
            target_df['_index'] = range(len(target_df))
            source_join_cols = ['_index']
            target_join_cols = ['_index']
        
        # Prepare DataFrames with mapped columns
        source_compare_df = source_df[source_cols].copy()
        target_compare_df = target_df[target_cols].copy()
        
        # Convert all columns to string type to avoid type comparison issues
        source_compare_df = source_compare_df.astype(str)
        target_compare_df = target_compare_df.astype(str)
        
        # Rename target columns to match source for proper comparison
        target_compare_df.columns = source_cols
        
        # Initialize comparison results
        comparison_results = {
            'total_rows_source': len(source_compare_df),
            'total_rows_target': len(target_compare_df),
            'matching_rows': 0,
            'source_only_rows': 0,
            'target_only_rows': 0,
            'differing_rows': 0,
            'source_only_df': pd.DataFrame(),
            'target_only_df': pd.DataFrame(),
            'diff_df': pd.DataFrame()
        }
        
        try:
            # Prepare join keys with type checking
            valid_join_keys = []
            
            # Default to index-based comparison
            source_compare_df['_index'] = range(len(source_compare_df))
            target_compare_df['_index'] = range(len(target_compare_df))
            valid_join_keys = ['_index']
            
            # Try to use provided join keys if available
            if join_keys:
                try:
                    # Convert join_keys to list if it's not already
                    join_key_list = list(join_keys) if isinstance(join_keys, (list, tuple)) else []
                    
                    temp_join_keys = []
                    for key in join_key_list:
                        if isinstance(key, (list, tuple)) and len(key) == 2:
                            source_key = str(key[0])
                            if source_key in source_compare_df.columns:
                                # Convert columns to string and handle nulls
                                source_compare_df[source_key] = source_compare_df[source_key].fillna('').astype(str)
                                target_compare_df[source_key] = target_compare_df[source_key].fillna('').astype(str)
                                temp_join_keys.append(source_key)
                    
                    # Only use custom join keys if valid ones were found
                    if temp_join_keys:
                        valid_join_keys = temp_join_keys
                        # Remove index columns if we're using custom join keys
                        if '_index' in source_compare_df.columns:
                            source_compare_df.drop('_index', axis=1, inplace=True)
                        if '_index' in target_compare_df.columns:
                            target_compare_df.drop('_index', axis=1, inplace=True)
                except Exception as e:
                    print(f"Warning: Error processing join keys, falling back to index-based comparison: {str(e)}")
            
            try:
                # Perform merge to find matching and non-matching rows
                merged_df = pd.merge(
                    source_compare_df,
                    target_compare_df,
                    on=valid_join_keys,
                    how='outer',
                    indicator=True
                )
            except Exception as e:
                raise ValueError(f"Error performing merge operation: {str(e)}")
            
            # Calculate comparison statistics
            comparison_results['source_only_df'] = merged_df[merged_df['_merge'] == 'left_only']
            comparison_results['target_only_df'] = merged_df[merged_df['_merge'] == 'right_only']
            matching_rows = merged_df[merged_df['_merge'] == 'both']
            
            comparison_results['source_only_rows'] = len(comparison_results['source_only_df'])
            comparison_results['target_only_rows'] = len(comparison_results['target_only_df'])
            comparison_results['matching_rows'] = len(matching_rows)
            
            # Calculate match percentage
            total_rows = len(source_compare_df)
            match_percentage = (comparison_results['matching_rows'] / total_rows * 100) if total_rows > 0 else 0
            
            # Generate summary
            summary = f"""
            Comparison Summary:
            - Total Rows in Source: {comparison_results['total_rows_source']}
            - Total Rows in Target: {comparison_results['total_rows_target']}
            - Matching Rows: {comparison_results['matching_rows']}
            - Source Only Rows: {comparison_results['source_only_rows']}
            - Target Only Rows: {comparison_results['target_only_rows']}
            - Match Percentage: {match_percentage:.2f}%
            """
            
            # Generate column statistics
            stats_data = []
            for col in source_compare_df.columns:
                if col != '_index':  # Skip index column if it exists
                    source_stats = {
                        'Column': col,
                        'Source_Count': len(source_compare_df[col].dropna()),
                        'Source_Unique': source_compare_df[col].nunique(),
                        'Target_Count': len(target_compare_df[col].dropna()),
                        'Target_Unique': target_compare_df[col].nunique(),
                        'Match_Status': 'Match' if source_compare_df[col].equals(target_compare_df[col]) else 'Mismatch'
                    }
                    stats_data.append(source_stats)
            
            col_stats = pd.DataFrame(stats_data)
            col_stats_html = col_stats.to_html(classes='table table-striped', index=False)
            
        except Exception as e:
            raise Exception(f"Error performing comparison: {str(e)}")
            
        # Generate HTML report
        try:
            # Generate HTML for source-only records
            source_only_html = comparison_results['source_only_df'].to_html(
                classes='table table-striped', 
                index=False
            ) if not comparison_results['source_only_df'].empty else '<p>No records unique to source</p>'
            
            # Generate HTML for target-only records
            target_only_html = comparison_results['target_only_df'].to_html(
                classes='table table-striped', 
                index=False
            ) if not comparison_results['target_only_df'].empty else '<p>No records unique to target</p>'
            
            # Write the HTML report
            with open(output_path, 'w') as f:
                f.write(f"""
                <!DOCTYPE html>
                <html lang="en">
                <head>
                    <title>DataCompy Comparison Report</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 20px; }}
                        .report {{ max-width: 1200px; margin: 0 auto; }}
                        .section {{ margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }}
                        .match {{ color: green; }}
                        .mismatch {{ color: red; }}
                        table {{ border-collapse: collapse; width: 100%; }}
                        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                        th {{ background-color: #f5f5f5; }}
                    </style>
                </head>
                <body>
                    <div class="report">
                        <h1>DataCompy Comparison Report</h1>
                        <div class="section">
                            <h2>Summary</h2>
                            <pre>{summary}</pre>
                        </div>
                        <div class="section">
                            <h2>Detailed Statistics</h2>
                        <h3>Matches</h3>
                        <div class="match">
                            <p>Number of rows match: {comparison_results['matching_rows']}</p>
                            <p>Number of columns match: {len(source_cols)}</p>
                        </div>
                        <h3>Mismatches</h3>
                        <div class="mismatch">
                            <p>Rows only in Source: {comparison_results['source_only_rows']}</p>
                            <p>Rows only in Target: {comparison_results['target_only_rows']}</p>
                        </div>
                        </div>
                        <div class="section">
                            <h2>Column Statistics</h2>
                            {col_stats_html}
                        </div>
                        <div class="section">
                            <h2>Mismatched Records</h2>
                            <h3>Records only in Source:</h3>
                            {source_only_html}
                            <h3>Records only in Target:</h3>
                            {target_only_html}
                        </div>
                    </div>
                </body>
                </html>
                """)
        except Exception as e:
            # Create a simple error report if something goes wrong
            with open(output_path, 'w') as f:
                f.write(f"""
                <html>
                <head><title>Comparison Error Report</title></head>
                <body>
                    <h1>Error Generating Comparison Report</h1>
                    <p>An error occurred while generating the comparison report: {str(e)}</p>
                </body>
                </html>
                """)
            raise Exception(f"Error generating DataCompy report: {str(e)}")
