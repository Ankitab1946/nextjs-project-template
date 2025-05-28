"""Module for generating comparison reports"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
import os
import ydata_profiling
import datacompy
import openpyxl
from openpyxl.styles import PatternFill
import zipfile

def generate_comparison_report(source_df: pd.DataFrame, target_df: pd.DataFrame,
                           column_mapping: Dict[str, str], join_keys: Optional[List[str]],
                           output_path: str) -> None:
    """Generate pandas-based comparison report"""
    try:
        # Get source and target columns
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        # Select relevant columns and copy DataFrames
        source_compare_df = source_df[source_cols].copy()
        target_compare_df = target_df[target_cols].copy()
        
        # Rename target columns to match source
        target_compare_df.columns = source_cols
        
        # Convert all columns to string and handle nulls
        for col in source_cols:
            source_compare_df[col] = source_compare_df[col].fillna('').astype(str)
            target_compare_df[col] = target_compare_df[col].fillna('').astype(str)
        
        # Perform merge
        if isinstance(join_keys, (list, tuple)) and join_keys:
            # Use provided join keys if they exist in source columns
            valid_keys = [k for k in join_keys if k in source_cols]
            if valid_keys:
                merged_df = pd.merge(
                    source_compare_df,
                    target_compare_df,
                    on=valid_keys,
                    how='outer',
                    indicator=True
                )
            else:
                # Fall back to index-based comparison if no valid join keys
                merged_df = pd.merge(
                    source_compare_df,
                    target_compare_df,
                    left_index=True,
                    right_index=True,
                    how='outer',
                    indicator=True
                )
        else:
            # Use index-based comparison if no join keys provided
            merged_df = pd.merge(
                source_compare_df,
                target_compare_df,
                left_index=True,
                right_index=True,
                how='outer',
                indicator=True
            )
        
        # Calculate statistics
        stats = {
            'total_rows_source': len(source_compare_df),
            'total_rows_target': len(target_compare_df),
            'matching_rows': len(merged_df[merged_df['_merge'] == 'both']),
            'source_only_rows': len(merged_df[merged_df['_merge'] == 'left_only']),
            'target_only_rows': len(merged_df[merged_df['_merge'] == 'right_only'])
        }
        
        # Calculate match percentage
        total_rows = len(source_compare_df)
        match_percentage = (stats['matching_rows'] / total_rows * 100) if total_rows > 0 else 0
        
        # Generate summary
        summary = f"""
        Comparison Summary:
        - Total Rows in Source: {stats['total_rows_source']}
        - Total Rows in Target: {stats['total_rows_target']}
        - Matching Rows: {stats['matching_rows']}
        - Source Only Rows: {stats['source_only_rows']}
        - Target Only Rows: {stats['target_only_rows']}
        - Match Percentage: {match_percentage:.2f}%
        """
        
        # Generate column statistics
        col_stats = []
        for col in source_cols:
            col_stats.append({
                'Column': col,
                'Source_Count': len(source_compare_df[col].dropna()),
                'Source_Unique': source_compare_df[col].nunique(),
                'Target_Count': len(target_compare_df[col].dropna()),
                'Target_Unique': target_compare_df[col].nunique(),
                'Match_Status': 'Match' if source_compare_df[col].equals(target_compare_df[col]) else 'Mismatch'
            })
        
        col_stats_df = pd.DataFrame(col_stats)
        col_stats_html = col_stats_df.to_html(classes='table table-striped', index=False)
        
        # Generate HTML for mismatched records
        source_only = merged_df[merged_df['_merge'] == 'left_only']
        target_only = merged_df[merged_df['_merge'] == 'right_only']
        
        source_only_html = source_only.to_html(classes='table table-striped', index=False) if not source_only.empty else '<p>No records unique to source</p>'
        target_only_html = target_only.to_html(classes='table table-striped', index=False) if not target_only.empty else '<p>No records unique to target</p>'
        
        # Write HTML report
        with open(output_path, 'w') as f:
            f.write(f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <title>Data Comparison Report</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .report {{ max-width: 1200px; margin: 0 auto; }}
                    .section {{ margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }}
                    .match {{ color: green; }}
                    .mismatch {{ color: red; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f5f5f5; }}
                </style>
            </head>
            <body>
                <div class="report">
                    <h1>Data Comparison Report</h1>
                    <div class="section">
                        <h2>Summary</h2>
                        <pre>{summary}</pre>
                    </div>
                    <div class="section">
                        <h2>Column Statistics</h2>
                        {col_stats_html}
                    </div>
                    <div class="section">
                        <h2>Mismatched Records</h2>
                        <h3>Records only in Source:</h3>
                        {source_only_html}
                        <h3>Records only in Target:</h3>
                        {target_only_html}
                    </div>
                </div>
            </body>
            </html>
            """)
            
    except Exception as e:
        # Generate error report
        error_html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head><title>Comparison Error Report</title></head>
        <body>
            <h1>Error Generating Comparison Report</h1>
            <p>An error occurred: {str(e)}</p>
            <p>Please check that:</p>
            <ul>
                <li>All required columns exist in both datasets</li>
                <li>Join keys are valid column names</li>
                <li>Data types are compatible for comparison</li>
            </ul>
        </body>
        </html>
        """
        with open(output_path, 'w') as f:
            f.write(error_html)
        raise Exception(f"Error generating comparison report: {str(e)}")

def generate_ydata_profile_reports(source_df: pd.DataFrame, target_df: pd.DataFrame, timestamp: str) -> Dict[str, str]:
    """Generate ydata profiling reports for source, target, and comparison"""
    try:
        source_profile = source_df.profile_report(title="Source Data Profile")
        target_profile = target_df.profile_report(title="Target Data Profile")
        comparison_profile = source_df.profile_report(title="Comparison Data Profile")  # Placeholder for actual comparison
        
        reports_dir = os.path.join("reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        source_path = os.path.join(reports_dir, f"SourceProfile_{timestamp}.html")
        target_path = os.path.join(reports_dir, f"TargetProfile_{timestamp}.html")
        comparison_path = os.path.join(reports_dir, f"ComparisonProfile_{timestamp}.html")
        
        source_profile.to_file(source_path)
        target_profile.to_file(target_path)
        comparison_profile.to_file(comparison_path)
        
        return {
            "source": source_path,
            "target": target_path,
            "comparison": comparison_path
        }
    except Exception as e:
        raise Exception(f"Error generating ydata profiling reports: {str(e)}")

def generate_datacompy_report(source_df: pd.DataFrame, target_df: pd.DataFrame, column_mapping: Dict[str, str], join_keys: Optional[List[str]], timestamp: str) -> str:
    """Generate DataCompy comparison report"""
    try:
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        source_compare_df = source_df[source_cols].copy()
        target_compare_df = target_df[target_cols].copy()
        
        target_compare_df.columns = source_cols
        
        join_cols = join_keys if join_keys else []
        
        compare = datacompy.Compare(
            source_compare_df,
            target_compare_df,
            join_columns=join_cols,
            abs_tol=0,
            rel_tol=0,
            df1_name='Source',
            df2_name='Target'
        )
        
        report = compare.report()
        
        reports_dir = os.path.join("reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        report_path = os.path.join(reports_dir, f"DataCompyReport_{timestamp}.txt")
        with open(report_path, 'w') as f:
            f.write(report)
        
        return report_path
    except Exception as e:
        raise Exception(f"Error generating DataCompy report: {str(e)}")

def generate_regression_report(source_df: pd.DataFrame, target_df: pd.DataFrame, column_mapping: Dict[str, str], output_path: str) -> None:
    """Generate Excel regression report with aggregation, count, distinct checks"""
    try:
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        source_compare_df = source_df[source_cols].copy()
        target_compare_df = target_df[target_cols].copy()
        
        target_compare_df.columns = source_cols
        
        wb = openpyxl.Workbook()
        
        # AggregationCheck sheet
        agg_sheet = wb.active
        agg_sheet.title = "AggregationCheck"
        agg_sheet.append(["Column", "Source Sum", "Target Sum", "Result"])
        
        for col in source_cols:
            if pd.api.types.is_numeric_dtype(source_compare_df[col]):
                source_sum = source_compare_df[col].sum()
                target_sum = target_compare_df[col].sum()
                result = "PASS" if np.isclose(source_sum, target_sum) else "FAIL"
                agg_sheet.append([col, source_sum, target_sum, result])
        
        # Color coding
        for row in agg_sheet.iter_rows(min_row=2, max_col=4):
            cell = row[3]
            if cell.value == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Count Check sheet
        count_sheet = wb.create_sheet("CountCheck")
        count_sheet.append(["Source File Name", "Source Path", "Data Count", "Comparison"])
        
        source_count = len(source_compare_df)
        target_count = len(target_compare_df)
        result = "PASS" if source_count == target_count else "FAIL"
        count_sheet.append(["Source", "N/A", source_count, result])
        
        for row in count_sheet.iter_rows(min_row=2, max_col=4):
            cell = row[3]
            if cell.value == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # DistinctCheck sheet
        distinct_sheet = wb.create_sheet("DistinctCheck")
        distinct_sheet.append(["Column", "Source Distinct Count", "Target Distinct Count", "Result"])
        
        for col in source_cols:
            if not pd.api.types.is_numeric_dtype(source_compare_df[col]):
                source_distinct = source_compare_df[col].nunique()
                target_distinct = target_compare_df[col].nunique()
                result = "PASS" if source_distinct == target_distinct else "FAIL"
                distinct_sheet.append([col, source_distinct, target_distinct, result])
        
        for row in distinct_sheet.iter_rows(min_row=2, max_col=4):
            cell = row[3]
            if cell.value == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        wb.save(output_path)
        
    except Exception as e:
        raise Exception(f"Error generating regression report: {str(e)}")

def generate_side_by_side_report(source_df: pd.DataFrame, target_df: pd.DataFrame, column_mapping: Dict[str, str], join_columns: Optional[List[str]], output_path: str) -> bool:
    """Generate side-by-side difference Excel report"""
    try:
        source_cols = list(column_mapping.keys())
        target_cols = [column_mapping[col] for col in source_cols]
        
        source_compare_df = source_df[source_cols].copy()
        target_compare_df = target_df[target_cols].copy()
        
        target_compare_df.columns = source_cols
        
        if join_columns:
            merged_df = pd.merge(
                source_compare_df,
                target_compare_df,
                on=join_columns,
                how='outer',
                suffixes=('_source', '_target'),
                indicator=True
            )
        else:
            merged_df = pd.merge(
                source_compare_df,
                target_compare_df,
                left_index=True,
                right_index=True,
                how='outer',
                suffixes=('_source', '_target'),
                indicator=True
            )
        
        # Identify differences
        diff_rows = []
        for idx, row in merged_df.iterrows():
            for col in source_cols:
                source_val = row.get(f"{col}_source", None)
                target_val = row.get(f"{col}_target", None)
                if pd.isna(source_val) and pd.isna(target_val):
                    continue
                if source_val != target_val:
                    diff_rows.append(row)
                    break
        
        has_differences = len(diff_rows) > 0
        
        # Write to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if has_differences:
                diff_df = pd.DataFrame(diff_rows)
                diff_df.to_excel(writer, index=False, sheet_name='Differences')
            else:
                # Write a sheet with message "There is No Differences found"
                df_empty = pd.DataFrame({"Message": ["There is No Differences found"]})
                df_empty.to_excel(writer, index=False, sheet_name='Differences')
        
        return has_differences
    except Exception as e:
        raise Exception(f"Error generating side-by-side difference report: {str(e)}")

def zip_reports(report_paths: List[str], zip_path: str) -> None:
    """Zip multiple report files into a single archive"""
    try:
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file_path in report_paths:
                if os.path.exists(file_path):
                    zipf.write(file_path, os.path.basename(file_path))
    except Exception as e:
        raise Exception(f"Error creating zip archive: {str(e)}")
